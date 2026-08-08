"""
══════════════════════════════════════════════════════════
 NutriSmart AI — Load Test Report Generator
 Reads Locust CSV output and generates a rich Excel report
══════════════════════════════════════════════════════════

HOW TO USE
──────────
  1. Run Locust in headless mode first (saves CSV files):
       cd load-tests
       locust -f locust/locustfile.py \
           --host=http://127.0.0.1:8000 \
           --users=100 --spawn-rate=10 \
           --run-time=1m --headless \
           --csv=results/baseline

  2. Then run this report generator:
       python load-tests/generate_report.py

  3. Open: load-tests/results/Load_Test_Report.xlsx
"""

import os
import csv
import sys
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
os.makedirs(RESULTS_DIR, exist_ok=True)

STATS_CSV    = os.path.join(RESULTS_DIR, "baseline_stats.csv")
HISTORY_CSV  = os.path.join(RESULTS_DIR, "baseline_history.csv")
FAILURES_CSV = os.path.join(RESULTS_DIR, "baseline_failures.csv")
OUTPUT_XLSX  = os.path.join(RESULTS_DIR, "Load_Test_Report.xlsx")

# ── Styles ─────────────────────────────────────────────
HEADER_FILL  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP    = Alignment(vertical="top", wrap_text=True)


def style_header(ws, headers, widths=None):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill    = HEADER_FILL
        cell.font    = HEADER_FONT
        cell.alignment = CENTER
        cell.border  = THIN_BORDER
    ws.row_dimensions[1].height = 28
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def border_row(ws, row_idx, col_count):
    for c in range(1, col_count + 1):
        ws.cell(row=row_idx, column=c).border = THIN_BORDER
        ws.cell(row=row_idx, column=c).alignment = CENTER


def color_cell_by_value(cell, value, good_threshold, warn_threshold, lower_is_better=True):
    """Green / Orange / Red coloring based on performance thresholds."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return
    if lower_is_better:
        if v <= good_threshold:
            cell.fill = GREEN_FILL
        elif v <= warn_threshold:
            cell.fill = ORANGE_FILL
        else:
            cell.fill = RED_FILL
    else:
        if v >= good_threshold:
            cell.fill = GREEN_FILL
        elif v >= warn_threshold:
            cell.fill = ORANGE_FILL
        else:
            cell.fill = RED_FILL


def read_csv(path):
    if not os.path.exists(path):
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []
    return rows, fieldnames


# ─────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────
def build_report():
    stats_rows, _    = read_csv(STATS_CSV)
    history_rows, _  = read_csv(HISTORY_CSV)
    failures_rows, _ = read_csv(FAILURES_CSV)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════
    # SHEET 1 — Test Summary (Executive View)
    # ══════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Test Summary"
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 30

    # Title block
    ws_summary["A1"] = "NutriSmart AI — Baseline Load Test Report"
    ws_summary["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"].alignment = CENTER
    ws_summary.row_dimensions[1].height = 35

    ws_summary["A2"] = "Generated"
    ws_summary["B2"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A3"] = "Target URL"
    ws_summary["B3"] = "http://127.0.0.1:8000"
    ws_summary["A4"] = "Test Duration"
    ws_summary["B4"] = "60 seconds"
    ws_summary["A5"] = "Virtual Users"
    ws_summary["B5"] = "100 VUs"
    ws_summary["A6"] = "Spawn Rate"
    ws_summary["B6"] = "10 users/second"
    ws_summary["A7"] = "Tool"
    ws_summary["B7"] = "Locust + k6"

    for row in ws_summary.iter_rows(min_row=2, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.border    = THIN_BORDER
            cell.alignment = CENTER

    for r in [2, 3, 4, 5, 6, 7]:
        ws_summary.cell(row=r, column=1).font = Font(bold=True)

    # Aggregated stats from CSV (if available)
    ws_summary["A9"] = "Metric"
    ws_summary["B9"] = "Value"
    ws_summary["A9"].fill = HEADER_FILL
    ws_summary["A9"].font = HEADER_FONT
    ws_summary["A9"].alignment = CENTER
    ws_summary["B9"].fill = HEADER_FILL
    ws_summary["B9"].font = HEADER_FONT
    ws_summary["B9"].alignment = CENTER

    agg_row = next((r for r in stats_rows if r.get("Type") == "None" or r.get("Name") == "Aggregated"), None)
    summary_data = {}
    if agg_row:
        summary_data = {
            "Total Requests":       agg_row.get("Request Count", "N/A"),
            "Total Failures":       agg_row.get("Failure Count", "N/A"),
            "Requests / Second":    agg_row.get("Requests/s", "N/A"),
            "Avg Response Time":    f"{agg_row.get('Average Response Time', 'N/A')} ms",
            "Min Response Time":    f"{agg_row.get('Min Response Time', 'N/A')} ms",
            "Max Response Time":    f"{agg_row.get('Max Response Time', 'N/A')} ms",
            "Median Response Time": f"{agg_row.get('Median Response Time', 'N/A')} ms",
            "90th Percentile":      f"{agg_row.get('90%', 'N/A')} ms",
            "95th Percentile":      f"{agg_row.get('95%', 'N/A')} ms",
            "99th Percentile":      f"{agg_row.get('99%', 'N/A')} ms",
            "Failure Rate":         f"{round(float(agg_row.get('Failure Count',0) or 0) / max(float(agg_row.get('Request Count',1) or 1), 1) * 100, 2)} %",
        }
    else:
        summary_data = {
            "Total Requests":    "Run Locust first to generate CSV data",
            "Requests / Second": "—",
            "Avg Response Time": "—",
            "Min Response Time": "—",
            "Max Response Time": "—",
            "Failure Rate":      "—",
        }

    for i, (k, v) in enumerate(summary_data.items(), 10):
        ws_summary.cell(row=i, column=1).value = k
        ws_summary.cell(row=i, column=1).font  = Font(bold=True)
        ws_summary.cell(row=i, column=2).value = v
        ws_summary.cell(row=i, column=1).border = THIN_BORDER
        ws_summary.cell(row=i, column=2).border = THIN_BORDER
        ws_summary.cell(row=i, column=1).alignment = CENTER
        ws_summary.cell(row=i, column=2).alignment = CENTER

    # ══════════════════════════════════════════════════
    # SHEET 2 — Per Endpoint Stats
    # ══════════════════════════════════════════════════
    ws_stats = wb.create_sheet("Per Endpoint Stats")
    headers2 = [
        "Method", "Endpoint", "# Requests", "# Failures", "Failure %",
        "Avg (ms)", "Min (ms)", "Max (ms)", "Median (ms)",
        "90th % (ms)", "95th % (ms)", "99th % (ms)", "Req/s"
    ]
    style_header(ws_stats, headers2, [10, 40, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 10])

    if stats_rows:
        for i, row in enumerate(stats_rows, 2):
            reqs   = int(row.get("Request Count", 0) or 0)
            fails  = int(row.get("Failure Count", 0) or 0)
            fail_p = round(fails / max(reqs, 1) * 100, 2)
            data = [
                row.get("Type", ""),
                row.get("Name", ""),
                reqs,
                fails,
                f"{fail_p}%",
                row.get("Average Response Time", ""),
                row.get("Min Response Time", ""),
                row.get("Max Response Time", ""),
                row.get("Median Response Time", ""),
                row.get("90%", ""),
                row.get("95%", ""),
                row.get("99%", ""),
                row.get("Requests/s", ""),
            ]
            ws_stats.append(data)
            border_row(ws_stats, i, len(headers2))

            # Color avg response time cell
            avg_cell = ws_stats.cell(row=i, column=6)
            color_cell_by_value(avg_cell, row.get("Average Response Time", 0), 500, 1500)

            # Color failure % cell
            fail_cell = ws_stats.cell(row=i, column=5)
            color_cell_by_value(fail_cell, fail_p, 1, 5)
    else:
        ws_stats.append(["No data yet — run Locust first to populate this sheet."])

    # ══════════════════════════════════════════════════
    # SHEET 3 — Response Time Distribution
    # ══════════════════════════════════════════════════
    ws_dist = wb.create_sheet("Response Time Distribution")
    ws_dist["A1"] = "Response Time Benchmark Guide"
    ws_dist["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws_dist.merge_cells("A1:C1")

    benchmarks = [
        ("Response Time Range", "User Experience", "Status"),
        ("< 100 ms",   "Instant — feels instantaneous",         "Excellent"),
        ("100-300 ms", "Fast — user notices no lag",            "Good"),
        ("300-500 ms", "Acceptable — slight perceived delay",   "Acceptable"),
        ("500-1000 ms","Slow — user notices delay",             "Warning"),
        ("1000-2000 ms","Very Slow — frustrating",              "Poor"),
        ("> 2000 ms",  "Unacceptable — users abandon the page", "Critical"),
    ]
    for i, row in enumerate(benchmarks, 3):
        ws_dist.append(list(row))
        for c in range(1, 4):
            ws_dist.cell(row=i, column=c).border = THIN_BORDER
            ws_dist.cell(row=i, column=c).alignment = CENTER
        if i == 3:
            for c in range(1, 4):
                ws_dist.cell(row=i, column=c).fill = HEADER_FILL
                ws_dist.cell(row=i, column=c).font = HEADER_FONT
        else:
            status = row[2]
            fill = GREEN_FILL if status == "Excellent" else \
                   GREEN_FILL if status == "Good" else \
                   ORANGE_FILL if status == "Acceptable" else \
                   ORANGE_FILL if status == "Warning" else RED_FILL
            for c in range(1, 4):
                ws_dist.cell(row=i, column=c).fill = fill

    ws_dist.column_dimensions["A"].width = 20
    ws_dist.column_dimensions["B"].width = 40
    ws_dist.column_dimensions["C"].width = 15

    ws_dist["A11"] = "SLA Thresholds (this test)"
    ws_dist["A11"].font = Font(bold=True, size=12, color="1F3864")
    thresholds = [
        ("Threshold", "Target Value", "Pass Condition"),
        ("95th Percentile", "< 2000 ms", "All requests p95 under 2 seconds"),
        ("Average Response", "< 500 ms",  "Average must stay under 500ms"),
        ("Error Rate", "< 5%",       "Less than 5% of requests fail"),
        ("Login p95", "< 1500 ms", "Login endpoint 95th percentile"),
        ("Recipe Engine p95", "< 3000 ms", "Recipe suggestion 95th percentile"),
    ]
    for i, row in enumerate(thresholds, 12):
        ws_dist.append(list(row))
        for c in range(1, 4):
            ws_dist.cell(row=i, column=c).border = THIN_BORDER
            ws_dist.cell(row=i, column=c).alignment = CENTER
        if i == 12:
            for c in range(1, 4):
                ws_dist.cell(row=i, column=c).fill = HEADER_FILL
                ws_dist.cell(row=i, column=c).font = HEADER_FONT

    # ══════════════════════════════════════════════════
    # SHEET 4 — Failures
    # ══════════════════════════════════════════════════
    ws_fail = wb.create_sheet("Failures")
    fail_headers = ["Method", "Endpoint", "Error", "# Occurrences"]
    style_header(ws_fail, fail_headers, [10, 40, 60, 15])

    if failures_rows:
        for i, row in enumerate(failures_rows, 2):
            ws_fail.append([
                row.get("Method", ""),
                row.get("Name", ""),
                row.get("Error", ""),
                row.get("Occurrences", ""),
            ])
            border_row(ws_fail, i, 4)
            for c in range(1, 5):
                ws_fail.cell(row=i, column=c).fill = RED_FILL
    else:
        ws_fail["A2"] = "No failures recorded OR run Locust first to populate."
        ws_fail["A2"].font = Font(italic=True, color="4CAF50")

    # ══════════════════════════════════════════════════
    # SHEET 5 — How To Run Instructions
    # ══════════════════════════════════════════════════
    ws_how = wb.create_sheet("How To Run")
    instructions = [
        ("Step", "Command / Action"),
        ("1. Install k6 (Windows)",      "winget install k6  OR  choco install k6"),
        ("2. Run k6 Baseline Test",       "k6 run load-tests/k6/baseline-test.js"),
        ("3. Run k6 with JSON output",    "k6 run --out json=load-tests/results/k6-result.json load-tests/k6/baseline-test.js"),
        ("4. Install Locust",             "pip install locust"),
        ("5. Run Locust with Web UI",     "locust -f load-tests/locust/locustfile.py --host=http://127.0.0.1:8000"),
        ("6. Open Locust UI",             "http://localhost:8089 → Users: 100, Spawn: 10, Duration: 1m"),
        ("7. Run Locust Headless",        "locust -f load-tests/locust/locustfile.py --host=http://127.0.0.1:8000 --users=100 --spawn-rate=10 --run-time=1m --headless --csv=load-tests/results/baseline"),
        ("8. Generate this Excel report", "python load-tests/generate_report.py"),
        ("9. Open report",                "load-tests/results/Load_Test_Report.xlsx"),
    ]
    style_header(ws_how, ["Step", "Command / Action"], [45, 90])
    for i, row in enumerate(instructions[1:], 2):
        ws_how.append(list(row))
        border_row(ws_how, i, 2)
        ws_how.cell(row=i, column=1).font = Font(bold=True)
        ws_how.row_dimensions[i].height = 22

    # ── Save ───────────────────────────────────────────
    wb.save(OUTPUT_XLSX)
    print(f"[OK] Load Test Report saved -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    build_report()
